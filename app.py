import requests
import pandas as pd
import time
from datetime import datetime
from fp.fp import FreeProxy
from urllib3.exceptions import MaxRetryError
from requests.exceptions import SSLError, ConnectionError
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from tkinter import ttk
import threading 

def calculate_miles(from_airport, to_airport, dep_date, ret_date, dep_class, ret_class, adults, ofws, children, infants, teenagers, by, travel_type, tier):
    url = f"https://www.emirates.com/service/ekl/loyalty/calculate-miles?airline=EK&origin={from_airport}&destination={to_airport}&cabin={dep_class}&journeyType=OW&tier={tier}"

    # Update the cookie dynamically with provided parameters
    cookie_value = f'ps={{"From":"{from_airport}","To":"{to_airport}","DepDate":"{dep_date}","RetDate":"{ret_date}","DepClass":"{dep_class}","RetClass":"{ret_class}","Adults":"{adults}","OFWs":"{ofws}","Child":"{children}","Infants":"{infants}","Teenager":"{teenagers}","By":"{by}","Type":"{travel_type}"}};'

    headers = {
        'accept': 'application/json, text/plain, */*',
        'accept-language': 'en-US,en;q=0.9',
        'cookie': cookie_value,
        'priority': 'u=1, i',
        'referer': 'https://www.emirates.com/ph/english/skywards/miles-calculator/',
        'sec-ch-ua': '"Chromium";v="128", "Not;A=Brand";v="24", "Google Chrome";v="128"',
        'sec-ch-ua-mobile': '?1',
        'sec-ch-ua-platform': '"Android"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'same-origin',
        'user-agent': 'Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 Mobile Safari/537.36'
    }

    response = requests.get(url, headers=headers, timeout=60, verify=True)
    return response.json()

def get_valid_proxy(url, max_retries=5):
    retries = 0
    proxy = None
    
    while retries < max_retries:
        try:
            # Get a random proxy
            proxy = FreeProxy(timeout=1, rand=True).get()
            print(f"Trying proxy: {proxy}")
            
            # Set up the proxies for requests
            proxies = {
                'http': proxy,
                'https': proxy,
            }
            
            # Attempt to make a request with the proxy
            response = requests.get(url, proxies=proxies, timeout=30, verify=True)  # SSL verification enabled
            response.raise_for_status()  # Raise an HTTPError if the status is 4xx/5xx
            
            # If successful, return the proxy
            print(f"Valid proxy found: {proxy}")
            return proxy
        
        except (SSLError, ConnectionError, MaxRetryError) as e:
            print(f"Proxy failed ({e}), retrying... ({retries + 1}/{max_retries})")
            retries += 1
            
    # If no valid proxy was found after retries
    raise Exception("Max retries exceeded. Could not find a valid proxy.")

def get_cabin_code(cabin_class):
    cabin_mapping = {
        'Economy': 'Y',
        'Business': 'J',
        'First': 'F',
        'Premium Economy': 'W'
    }
    return cabin_mapping.get(cabin_class, 'Unknown')

# Function to get cabin name from code
def format_cabin_name(cabin_code):
    cabin_mapping = {
        'Y': 'Economy',
        'J': 'Business',
        'F': 'First',
        'W': 'Premium Economy'
    }
    return cabin_mapping.get(cabin_code, 'Unknown')

# Function to handle fare formatting
def format_fare(cabin_formatted, fare):
    if cabin_formatted == "Premium Economy":
        return f"{fare.capitalize()}"
    return f"{cabin_formatted} {fare.capitalize()}"

# Function to process each row of the DataFrame
def process_row(row):
    # Access row values
    one_way_or_roundtrip = row['oneWayOrRoundtrip']
    flying_with = row['flyingWith']
    leaving_from = row['leavingFrom']
    going_to = row['goingTo']
    cabin_class = row['cabinClass']
    emirates_skywards_tier = row['emiratesSkywardsTier']
    now = datetime.now()
    input_date = now.strftime("%d%m%Y")

    print(f"{leaving_from} - {going_to} - {cabin_class} - {emirates_skywards_tier} - Scraping...")

    dep_class = get_cabin_code(cabin_class)
    ret_class = "7"
    dep_date = input_date
    ret_date = "151024"
    adults = "2"
    ofws = "0"
    children = "0"
    infants = "0"
    teenagers = "0"
    by = "0"
    travel_type = "0"

    # Get miles data
    miles_data = calculate_miles(leaving_from, going_to, dep_date, ret_date, dep_class, ret_class, adults, ofws, children, infants, teenagers, by, travel_type, emirates_skywards_tier.lower())
    if not miles_data:
        print(f"{leaving_from} - {going_to} - {cabin_class} - {emirates_skywards_tier} - No data found.")
        return None  # Skip if no data

    
    print(f"{leaving_from} - {going_to} - {cabin_class} - {emirates_skywards_tier} - Scraped successfully!")
    return process_miles_data(leaving_from, going_to, cabin_class, miles_data, emirates_skywards_tier)

# Function to process miles data
def process_miles_data(leaving_from, going_to, cabin_class, miles_data, tier):
    # Extract relevant data
    # origin = miles_data['getMilesFromCouchbase']['origin']
    origin = leaving_from
    # destination = miles_data['getMilesFromCouchbase']['destination']
    destination = going_to
    # cabin = miles_data['getMilesFromCouchbase']['cabin']
    cabin = cabin_class
    # journey_type = miles_data['getMilesFromCouchbase']['journeyType']

    if tier.lower() == 'blue':
        tier = 'skywards'
    else:
        tier = tier.lower()

    earn_miles = (miles_data.get('getMilesFromCouchbase', {})
                  .get('miles', {})
                  .get('earn', {})
                  .get(tier, None))

    # Define fares to iterate over
    fares = ['flexPlus', 'flex', 'saver', 'special']
    rows = []

    for fare in fares:
        if cabin == 'First' and fare in ['saver', 'special']:
            continue  # Skip invalid fare combinations

        skywards_miles, tier_miles = extract_miles(earn_miles, fare)

        # Format cabin and fare
        cabin_formatted = cabin
        formatted_fare = format_fare(cabin_formatted, fare)

        if tier.lower() == 'skywards':
            tier = 'blue'
        else:
            tier = tier.lower()

        row = {
            'Direction': 'One Way',
            'Airline': 'Emirates',
            'Leaving from': origin,
            'Going to': destination,
            'Cabin Class': cabin_formatted,
            'Skywards Tier': tier.capitalize(),
            'Fare': formatted_fare,
            'Skywards Miles': skywards_miles,
            'Tier Miles': tier_miles
        }
        rows.append(row)
    return rows

# Function to extract miles from the data
def extract_miles(earn_miles, fare):
    if not earn_miles:
        print(f"{fare}: Route does not exist or miles data is not available.")
        return 'None', 'None'
    
    skywards_miles = earn_miles.get(fare, {}).get('skywardsMiles', 'N/A')
    tier_miles = earn_miles.get(fare, {}).get('tierMiles', 'N/A')

    if str(skywards_miles).isdigit():
        skywards_miles = int(skywards_miles)

    if str(tier_miles).isdigit():
        tier_miles = int(tier_miles)

    return skywards_miles, tier_miles

def save_and_format_excel(dataframe, default_filename):
    """
    Save a DataFrame to an Excel file and apply formatting.
    
    :param dataframe: The DataFrame to save.
    :param file_path: Path to the output Excel file.
    """

     # Open a save file dialog with a default filename
    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", 
                                                filetypes=[("Excel files", "*.xlsx")],
                                                initialfile=default_filename)
    if save_path:
        # Save the processed DataFrame to the user-defined path
        # Save DataFrame to Excel file using pandas
        with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
            dataframe.to_excel(writer, index=False, sheet_name='Earn')
        
        # Load the workbook and select the active sheet
        wb = load_workbook(save_path)
        ws = wb.active

        # Bold the headers
        header_row = ws[1]  # Assumes headers are in the first row
        for cell in header_row:
            cell.font = Font(bold=True)

        # Align all cells to the left
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='left')

        # Autofit column widths (approximation)
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter  # Get the column letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2  # Add a little extra space
            ws.column_dimensions[column_letter].width = adjusted_width

        # Autofit row heights (approximation)
        for row in ws.iter_rows():
            max_height = 0
            for cell in row:
                if cell.value:
                    max_height = max(max_height, len(str(cell.value)) // 10)  # Rough estimate
            ws.row_dimensions[row[0].row].height = max_height * 15  # Adjust factor as needed

        # Save the workbook with changes
        wb.save(save_path)
        messagebox.showinfo("Success", f"File saved successfully: {save_path}")
    else:
        messagebox.showwarning("Warning", "Save cancelled.")
    

# Main function to handle file selection and processing
def select_file():
    global df
    # Open a file dialog to choose the Excel file
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    
    if file_path:
        try:
            # Create DataFrame from the selected Excel file
            df = pd.read_excel(file_path)
            messagebox.showinfo("Success", f"File loaded successfully: {file_path}")
            progress_bar['value'] = 0
            progress_label.config(text="File loaded. Ready to scrape.")
            scrape_button.config(state=tk.NORMAL)
        except Exception as e:
            messagebox.showerror("Error", f"Error loading file: {e}")
    else:
        messagebox.showwarning("Warning", "No file selected.")

def scrape_data():
    global df
    if df is None:
        messagebox.showwarning("Warning", "No file loaded. Please select an Excel file first.")
        return
    
    try:
        progress_bar['value'] = 0
        progress_label.config(text="Scraping is currently in progress...")
        
        # Get total number of rows
        total_rows = df.shape[0]
        if total_rows == 0:
            messagebox.showwarning("Warning", "No rows in the selected file.")
            return

        rows = []
        now = datetime.now()
        route_num = "6"
        route = "[RouteCode]"
        batch = "1"
        formatted_date = now.strftime("%m%d%Y_%H%M")

        # Get valid proxy
        # url = "https://www.emirates.com/ph/english/skywards/miles-calculator/"
        # proxy = get_valid_proxy(url, max_retries=10)
        
        # if proxy:
        #     proxies = {'http': proxy, 'https': proxy}
        # else:
        #     proxies = {}

        # Process each row in the DataFrame
        try:
            for index, row in df.iterrows():
                time.sleep(4)
                processed_rows = process_row(row)
                if processed_rows:
                    rows.extend(processed_rows)
                progress = ((index + 1) / total_rows) * 100
                progress_bar['value'] = progress
                root.update_idletasks()  # Update the GUI

        except Exception as e:
            print(f"Error processing row {index}: {e}")
        finally:
            # Save data to Excel
            progress_label.config(text="Scraping complete...")
            # global end_time
            end_time = time.time()
            execution_time = end_time - start_time
            
            if execution_time < 60:
                print(f"Execution time: {execution_time:.2f} seconds")
                progress_label.config(text=f"Execution time: {execution_time:.2f} seconds")
            elif execution_time < 3600:
                minutes = execution_time / 60
                print(f"Execution time: {minutes:.2f} minutes")
                progress_label.config(text=f"Execution time: {minutes:.2f} minutes")
            else:
                hours = execution_time / 3600
                print(f"Execution time: {hours:.2f} hours")
                progress_label.config(text=f"Execution time: {hours:.2f} hours")
            default_filename = f'{route_num} EKS_{route}_{formatted_date}_{batch}.xlsx'
            save_and_format_excel(pd.DataFrame(rows), default_filename)
    except Exception as e:
        messagebox.showerror("Error", f"Error processing or saving file: {e}")

# Function to start the scraping process in a new thread
def start_scraping_thread():
    scrape_thread = threading.Thread(target=scrape_data)
    scrape_thread.start()

# Main execution logic
def main():
    global root, scrape_button, progress_bar, progress_label, start_time

    # Set up the main tkinter window
    root = tk.Tk()
    root.title("EK Miles Scraper")

    # Set window size and background color
    root.geometry("600x300")
    root.configure(bg="white")

    start_time = time.time()

    # Add a label with custom font and colors
    label = tk.Label(root, text="EK Miles Scraper", font=("Arial", 24, "bold"), bg="white", fg="red")
    label.pack(pady=20)

    # Button to select the file
    select_button = tk.Button(root, text="Select File", command=select_file, width=20)
    select_button.pack(pady=10)

    # Initially, the scrape button is disabled
    scrape_button = tk.Button(root, text="Scrape", command=start_scraping_thread, state=tk.DISABLED, width=20)
    scrape_button.pack(pady=10)

    # Progress bar widget
    progress_bar = ttk.Progressbar(root, orient="horizontal", length=400, mode="determinate")
    progress_bar.pack(pady=20)

    # Progress label widget to display the status of the scraping
    progress_label = tk.Label(root, text="Waiting for file...")
    progress_label.pack(pady=10)

    root.mainloop()

if __name__ == "__main__":
    main()